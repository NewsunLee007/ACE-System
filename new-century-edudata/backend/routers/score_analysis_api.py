"""
成绩分析API模块
支持分层教学分析和成果发布
包含权限控制、数据分析、可视化、发布等功能
"""

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, func
from typing import List, Optional, Dict, Any
from pydantic import BaseModel
from datetime import datetime
import json
import uuid
import logging
import math
import statistics
import io
from types import SimpleNamespace

from openpyxl import Workbook
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
except ImportError:  # pragma: no cover - dependency may be absent in minimal test envs
    colors = None
    A4 = None
    getSampleStyleSheet = None
    SimpleDocTemplate = None
    Paragraph = None
    Spacer = None
    Table = None
    TableStyle = None
    pdfmetrics = None
    UnicodeCIDFont = None

from core.database import get_db, is_postgresql, is_sqlite
from core.security import get_current_user, has_permission_code, require_permission, require_permission_codes
from services.score_visibility_service import (
    fetch_score_visibility_settings,
    filter_rank_fields_by_visibility,
    resolve_visibility_for_user,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1/score-analysis", tags=["成绩分析"])


def _row_with_subject_scores(row: Any) -> Any:
    mapping = dict(row._mapping) if hasattr(row, "_mapping") else dict(vars(row))
    mapping["scores"] = json.dumps({
        "chinese": mapping.get("score_chinese"),
        "math": mapping.get("score_math"),
        "english": mapping.get("score_english"),
        "science": mapping.get("score_science"),
        "society": mapping.get("score_society"),
    }, ensure_ascii=False, default=str)
    return SimpleNamespace(**mapping)

DEFAULT_GRADE_LEVELS = ["7年级", "8年级", "9年级"]
SCORE_ANALYSIS_VIEW_PERMISSION_CODES = (
    "exam_admin",
    "grade_leader",
    "subject_leader",
)
SCORE_ANALYSIS_EXECUTE_PERMISSION_CODES = (
    "exam_admin",
    "grade_leader",
    "subject_leader",
)
LEGACY_ROLE_PERMISSION_CODES = {
    "super_admin": "sys_admin",
    "dean": "edu_admin",
    "school_leader": "edu_admin",
    "middle_manager": "exam_admin",
    "research_leader": "subject_leader",
    "prep_leader": "lesson_leader",
    "head_teacher": "headmaster",
    "subject_teacher": "teacher",
    "教务处主任": "edu_admin",
    "教务处主任/校领导": "edu_admin",
    "系统管理员": "sys_admin",
    "考务与学籍管理员": "exam_admin",
    "年段长": "grade_leader",
    "教研组长": "subject_leader",
    "备课组长": "lesson_leader",
    "班主任": "headmaster",
    "科任教师": "teacher",
}
RECIPIENT_TYPE_PERMISSION_CODES = {
    "school_leader": ("edu_admin", "sys_admin"),
    "middle_manager": ("exam_admin", "grade_leader"),
    "research_leader": ("subject_leader",),
    "prep_leader": ("lesson_leader",),
    "grade_leader": ("grade_leader",),
    "head_teacher": ("headmaster",),
    "teacher": ("teacher", "headmaster", "lesson_leader", "subject_leader", "grade_leader"),
}


# ============ Pydantic模型定义 ============

class LayerConfig(BaseModel):
    """层次配置"""
    code: str
    name: str
    description: Optional[str] = None


class ClassLayerSetting(BaseModel):
    """班级层次设定"""
    grade_level: str
    class_id: int
    class_name: str
    layer_code: str
    layer_name: str
    academic_year: str
    term: str
    description: Optional[str] = None


class AnalysisRequest(BaseModel):
    """分析请求"""
    exam_id: int
    grade_level: str
    analysis_type: str  # overall, layer_comparison, subject_analysis, student_progress, class_contrast
    analysis_scope: str  # all, layer_a, layer_b, layer_c


class AnalysisResult(BaseModel):
    """分析结果"""
    analysis_id: str
    exam_id: int
    exam_name: str
    grade_level: str
    analysis_type: str
    analysis_scope: str
    analysis_data: Dict[str, Any]
    created_by: int
    created_by_name: str
    created_at: str
    status: str


class PublicationRequest(BaseModel):
    """发布请求"""
    analysis_id: str
    title: str
    content_summary: Optional[str] = None
    recipient_types: List[str]  # school_leader, middle_manager, research_leader, etc.


class PublicationRecipient(BaseModel):
    """发布接收对象"""
    user_id: int
    user_name: str
    user_role: str
    read_status: str
    read_at: Optional[str] = None


class PublicationResult(BaseModel):
    """发布结果"""
    publication_id: str
    analysis_id: str
    exam_id: int
    exam_name: str
    grade_level: str
    title: str
    published_by: int
    published_by_name: str
    published_at: str
    recipient_types: List[str]
    recipient_count: int
    status: str


class BundleRefreshRequest(BaseModel):
    grade_level: Optional[str] = None


# ============ 权限检查函数 ============

def check_analysis_permission(current_user: dict, required_roles: List[str]) -> bool:
    """检查用户是否有成绩分析权限"""
    allowed_permission_codes = {
        LEGACY_ROLE_PERMISSION_CODES.get(role, role)
        for role in required_roles
        if role
    }
    return has_permission_code(current_user, allowed_permission_codes)


def _fetch_all_grade_levels(db: Optional[Session]) -> List[str]:
    if db is None:
        return DEFAULT_GRADE_LEVELS.copy()

    rows = db.execute(
        text("""
            SELECT DISTINCT grade_level
            FROM biz_exams
            WHERE grade_level IS NOT NULL AND grade_level <> ''
            ORDER BY grade_level
        """)
    ).fetchall()
    grades = [row.grade_level for row in rows if row.grade_level]
    return grades or DEFAULT_GRADE_LEVELS.copy()


def _fetch_user_relation_grades(current_user: dict, db: Optional[Session]) -> List[str]:
    if db is None or not current_user.get("id"):
        return []

    rows = db.execute(
        text("""
            SELECT DISTINCT grade_name
            FROM biz_teacher_class_rel
            WHERE teacher_id = :teacher_id
              AND grade_name IS NOT NULL
              AND grade_name <> ''
            ORDER BY grade_name
        """),
        {"teacher_id": current_user.get("id")}
    ).fetchall()
    return [row.grade_name for row in rows if row.grade_name]


def get_accessible_grades(current_user: dict, db: Optional[Session] = None) -> List[str]:
    """获取用户可访问的年级列表"""
    permission_code = current_user.get("permission_code")
    configured_grade = current_user.get("grade_level") or current_user.get("grade_name")

    if permission_code in {"sys_admin", "edu_admin", "exam_admin", "subject_leader"}:
        return _fetch_all_grade_levels(db)

    if permission_code == "grade_leader":
        if configured_grade:
            return [configured_grade]
        return _fetch_user_relation_grades(current_user, db)

    return []


def parse_class_id(class_name: Any) -> Optional[int]:
    text_value = str(class_name or "").strip()
    digits = "".join(ch for ch in text_value if ch.isdigit())
    if not digits:
        return None
    try:
        return int(digits[:4])
    except ValueError:
        return None


def build_score_subject_map(row, subjects: List[str]) -> Dict[str, Optional[float]]:
    subject_columns = {
        "语文": row.score_chinese,
        "数学": row.score_math,
        "英语": row.score_english,
        "科学": row.score_science,
        "社会": row.score_society,
    }
    configured_subjects = subjects or list(subject_columns.keys())
    return {
        subject: float(subject_columns[subject]) if subject in subject_columns and subject_columns[subject] is not None else None
        for subject in configured_subjects
        if subject in subject_columns
    }


def _add_in_condition(
    conditions: List[str],
    params: Dict[str, Any],
    field_name: str,
    values: List[str],
    prefix: str
) -> bool:
    cleaned_values = [value for value in values if value]
    if not cleaned_values:
        return False

    placeholders = []
    for index, value in enumerate(cleaned_values):
        param_name = f"{prefix}_{index}"
        params[param_name] = value
        placeholders.append(f":{param_name}")

    conditions.append(f"{field_name} IN ({', '.join(placeholders)})")
    return True


@router.get("/exams/{exam_id}/scores")
def get_exam_score_rows(
    exam_id: int,
    include_invalid: bool = Query(True, description="是否包含不参与统计/缺考成绩"),
    current_user: dict = Depends(require_permission_codes(*SCORE_ANALYSIS_VIEW_PERMISSION_CODES)),
    db: Session = Depends(get_db),
):
    """
    获取指定考试的原始成绩行。

    仅做只读取数，前端继续复用同一套成绩分析计算逻辑。
    """
    try:
        exam = db.execute(text("""
            SELECT id, exam_name, grade_level, subjects
            FROM biz_exams
            WHERE id = :exam_id
        """), {"exam_id": exam_id}).fetchone()
        if not exam:
            raise HTTPException(status_code=404, detail="考试不存在")

        accessible_grades = get_accessible_grades(current_user, db)
        if exam.grade_level and exam.grade_level not in accessible_grades:
            raise HTTPException(status_code=403, detail="无权访问该考试成绩")

        subjects = []
        if exam.subjects:
            try:
                subjects = json.loads(exam.subjects) if isinstance(exam.subjects, str) else exam.subjects
            except Exception:
                subjects = []
        if not isinstance(subjects, list):
            subjects = []

        conditions = ["s.exam_id = :exam_id"]
        params = {"exam_id": exam_id}
        if not include_invalid:
            conditions.append("s.is_included = 1")

        rows = db.execute(text(f"""
            SELECT
                s.id,
                s.exam_id,
                s.student_id,
                st.student_code,
                st.name AS student_name,
                s.exam_number,
                s.class_name,
                s.score_chinese,
                s.score_math,
                s.score_english,
                s.score_science,
                s.score_society,
                s.total_score,
                s.is_included,
                s.remarks,
                s.created_at,
                s.updated_at
            FROM biz_scores s
            LEFT JOIN biz_students st ON s.student_id = st.id
            WHERE {" AND ".join(conditions)}
            ORDER BY s.class_name, st.name, s.exam_number
        """), params).fetchall()

        score_rows = []
        for row in rows:
            score_rows.append({
                "id": row.id,
                "exam_id": row.exam_id,
                "student_id": row.student_id,
                "student_code": row.student_code or "",
                "student_name": row.student_name or "",
                "exam_number": row.exam_number or "",
                "class_id": parse_class_id(row.class_name),
                "class_name": row.class_name or "",
                "scores": build_score_subject_map(row, subjects),
                "total_score": float(row.total_score) if row.total_score is not None else None,
                "is_valid": row.is_included == 1,
                "is_included": row.is_included == 1,
                "remarks": row.remarks or "",
                "created_at": row.created_at.isoformat() if row.created_at else "",
                "updated_at": row.updated_at.isoformat() if row.updated_at else "",
            })

        return {
            "success": True,
            "exam_id": exam_id,
            "exam_name": exam.exam_name,
            "grade_level": exam.grade_level,
            "subjects": subjects,
            "total": len(score_rows),
            "scores": score_rows,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取考试成绩行失败: {e}")
        raise HTTPException(status_code=500, detail="获取考试成绩行失败")


# ============ 数据分析算法 ============

def calculate_basic_statistics(scores: List[float]) -> Dict[str, Any]:
    """计算基础统计指标"""
    if not scores:
        return {
            'count': 0,
            'mean': 0,
            'median': 0,
            'std': 0,
            'min': 0,
            'max': 0,
            'q1': 0,
            'q3': 0
        }
    
    sorted_scores = sorted(scores)
    n = len(sorted_scores)
    
    # 计算均值
    mean_val = sum(scores) / n
    
    # 计算中位数
    if n % 2 == 0:
        median_val = (sorted_scores[n//2 - 1] + sorted_scores[n//2]) / 2
    else:
        median_val = sorted_scores[n//2]
    
    # 计算标准差
    variance = sum((x - mean_val) ** 2 for x in scores) / n
    std_val = math.sqrt(variance)
    
    # 计算四分位数
    q1_val = sorted_scores[n // 4] if n >= 4 else sorted_scores[0]
    q3_val = sorted_scores[3 * n // 4] if n >= 4 else sorted_scores[-1]
    
    return {
        'count': n,
        'mean': round(mean_val, 2),
        'median': round(median_val, 2),
        'std': round(std_val, 2),
        'min': round(sorted_scores[0], 2),
        'max': round(sorted_scores[-1], 2),
        'q1': round(q1_val, 2),
        'q3': round(q3_val, 2)
    }


def calculate_score_distribution(scores: List[float], 
                                  thresholds: Dict[str, float]) -> Dict[str, int]:
    """计算成绩分布"""
    distribution = {
        'excellent': 0,  # 优秀
        'good': 0,       # 良好
        'pass': 0,       # 及格
        'fail': 0        # 不及格
    }
    
    for score in scores:
        if score >= thresholds.get('excellent', 90):
            distribution['excellent'] += 1
        elif score >= thresholds.get('good', 80):
            distribution['good'] += 1
        elif score >= thresholds.get('pass', 60):
            distribution['pass'] += 1
        else:
            distribution['fail'] += 1
    
    return distribution


def calculate_z_scores(scores: List[float], 
                        reference_mean: float, 
                        reference_std: float) -> List[float]:
    """计算Z分数"""
    if reference_std == 0:
        return [0.0] * len(scores)
    return [(score - reference_mean) / reference_std for score in scores]


def perform_t_test(group1_scores: List[float], 
                   group2_scores: List[float]) -> Dict[str, Any]:
    """执行T检验 (简化版，不使用scipy)"""
    if len(group1_scores) < 2 or len(group2_scores) < 2:
        return {'t_statistic': 0, 'p_value': 1, 'significant': False}
    
    n1, n2 = len(group1_scores), len(group2_scores)
    mean1, mean2 = sum(group1_scores) / n1, sum(group2_scores) / n2
    
    # 计算标准差
    var1 = sum((x - mean1) ** 2 for x in group1_scores) / (n1 - 1) if n1 > 1 else 0
    var2 = sum((x - mean2) ** 2 for x in group2_scores) / (n2 - 1) if n2 > 1 else 0
    
    # 计算合并标准差
    pooled_std = math.sqrt(((n1 - 1) * var1 + (n2 - 1) * var2) / (n1 + n2 - 2))
    
    if pooled_std == 0:
        return {'t_statistic': 0, 'p_value': 1, 'significant': False}
    
    # 计算t统计量
    t_stat = (mean1 - mean2) / (pooled_std * math.sqrt(1/n1 + 1/n2))
    
    # 简化的p值估计 (使用经验公式)
    df = n1 + n2 - 2
    p_value = min(1.0, max(0.001, 2 * (1 - min(1, abs(t_stat) / math.sqrt(df)))))
    
    return {
        't_statistic': round(t_stat, 4),
        'p_value': round(p_value, 4),
        'significant': p_value < 0.05
    }


# ============ 班级层次管理API ============

@router.get("/layers/config")
def get_layer_config(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取层次配置"""
    try:
        sql = "SELECT config_value FROM biz_analysis_configs WHERE config_key = 'layer_definitions'"
        result = db.execute(text(sql)).fetchone()
        
        if result:
            return {
                "success": True,
                "layers": json.loads(result.config_value)
            }
        
        # 默认配置
        default_layers = [
            {"code": "A", "name": "提高班", "description": "学业水平较高的班级"},
            {"code": "B", "name": "平行班", "description": "标准教学班级"},
            {"code": "C", "name": "基础班", "description": "需要加强基础的班级"}
        ]
        
        return {
            "success": True,
            "layers": default_layers
        }
        
    except Exception as e:
        logger.error(f"获取层次配置失败: {e}")
        raise HTTPException(status_code=500, detail="获取层次配置失败")


@router.post("/layers/setting")
def set_class_layer(
    setting: ClassLayerSetting,
    current_user: dict = Depends(require_permission("layer_manage")),
    db: Session = Depends(get_db)
):
    """设置班级层次"""
    try:
        # 检查是否已存在
        check_sql = """
            SELECT id FROM biz_class_layers 
            WHERE grade_level = :grade_level AND class_id = :class_id 
            AND academic_year = :academic_year AND term = :term
        """
        existing = db.execute(text(check_sql), {
            "grade_level": setting.grade_level,
            "class_id": setting.class_id,
            "academic_year": setting.academic_year,
            "term": setting.term
        }).fetchone()
        
        if existing:
            # 更新
            update_sql = """
                UPDATE biz_class_layers 
                SET layer_code = :layer_code, layer_name = :layer_name,
                    description = :description, created_by = :created_by
                WHERE id = :id
            """
            db.execute(text(update_sql), {
                "layer_code": setting.layer_code,
                "layer_name": setting.layer_name,
                "description": setting.description,
                "created_by": current_user["id"],
                "id": existing.id
            })
        else:
            # 插入
            insert_sql = """
                INSERT INTO biz_class_layers 
                (grade_level, class_id, class_name, layer_code, layer_name,
                 academic_year, term, description, created_by)
                VALUES 
                (:grade_level, :class_id, :class_name, :layer_code, :layer_name,
                 :academic_year, :term, :description, :created_by)
            """
            db.execute(text(insert_sql), {
                **setting.dict(),
                "created_by": current_user["id"]
            })
        
        db.commit()
        
        # 记录日志
        log_sql = """
            INSERT INTO biz_analysis_logs 
            (action_type, action_by, action_by_name, action_by_role, action_detail)
            VALUES 
            ('set_layer', :action_by, :action_by_name, :action_by_role, :action_detail)
        """
        db.execute(text(log_sql), {
            "action_by": current_user["id"],
            "action_by_name": current_user.get("real_name", current_user["username"]),
            "action_by_role": current_user.get("role_name", current_user.get("permission_code", "")),
            "action_detail": json.dumps(setting.dict(), ensure_ascii=False)
        })
        db.commit()
        
        return {"success": True, "message": "班级层次设置成功"}
        
    except Exception as e:
        logger.error(f"设置班级层次失败: {e}")
        raise HTTPException(status_code=500, detail="设置班级层次失败")


@router.get("/layers/list")
def get_class_layers(
    grade_level: Optional[str] = Query(None),
    academic_year: Optional[str] = Query(None),
    term: Optional[str] = Query(None),
    current_user: dict = Depends(require_permission_codes(*SCORE_ANALYSIS_VIEW_PERMISSION_CODES)),
    db: Session = Depends(get_db)
):
    """获取班级层次列表"""
    try:
        # 检查权限
        accessible_grades = get_accessible_grades(current_user, db)
        if grade_level and grade_level not in accessible_grades:
            raise HTTPException(status_code=403, detail="无权访问该年级数据")
        
        conditions = []
        params = {}
        
        if grade_level:
            conditions.append("grade_level = :grade_level")
            params["grade_level"] = grade_level
        else:
            if not _add_in_condition(conditions, params, "grade_level", accessible_grades, "grade"):
                return {"success": True, "layers": []}
        
        if academic_year:
            conditions.append("academic_year = :academic_year")
            params["academic_year"] = academic_year
        
        if term:
            conditions.append("term = :term")
            params["term"] = term
        
        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""
        
        sql = f"""
            SELECT * FROM biz_class_layers 
            {where_clause}
            ORDER BY grade_level, class_name
        """
        
        results = db.execute(text(sql), params).fetchall()
        
        layers = []
        for result in results:
            layers.append({
                "id": result.id,
                "grade_level": result.grade_level,
                "class_id": result.class_id,
                "class_name": result.class_name,
                "layer_code": result.layer_code,
                "layer_name": result.layer_name,
                "academic_year": result.academic_year,
                "term": result.term,
                "description": result.description
            })
        
        return {
            "success": True,
            "layers": layers
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取班级层次列表失败: {e}")
        raise HTTPException(status_code=500, detail="获取班级层次列表失败")


# ============ 成绩分析API ============

@router.post("/analyze")
def perform_analysis(
    request: AnalysisRequest,
    req: Request,
    current_user: dict = Depends(require_permission_codes(*SCORE_ANALYSIS_EXECUTE_PERMISSION_CODES)),
    db: Session = Depends(get_db)
):
    """
    执行成绩分析
    
    支持多种分析类型：
    - overall: 整体分析
    - layer_comparison: 层次对比分析
    - subject_analysis: 学科分析
    - student_progress: 学生进退步分析
    - class_contrast: 班级对比分析
    """
    try:
        # 检查权限
        if not check_analysis_permission(current_user, list(SCORE_ANALYSIS_EXECUTE_PERMISSION_CODES)):
            raise HTTPException(status_code=403, detail="无权进行成绩分析")
        
        accessible_grades = get_accessible_grades(current_user, db)
        if request.grade_level not in accessible_grades:
            raise HTTPException(status_code=403, detail="无权访问该年级数据")
        
        # 获取考试信息
        exam_sql = "SELECT exam_name, grade_level FROM biz_exams WHERE id = :exam_id"
        exam_result = db.execute(text(exam_sql), {"exam_id": request.exam_id}).fetchone()
        if not exam_result:
            raise HTTPException(status_code=404, detail="考试不存在")
        if exam_result.grade_level and exam_result.grade_level != request.grade_level:
            raise HTTPException(status_code=400, detail="考试年级与分析年级不一致")
        
        exam_name = exam_result.exam_name
        
        # 获取成绩数据
        scores_sql = """
            SELECT
                s.id,
                s.exam_id,
                s.student_id,
                st.name AS student_name,
                s.exam_number,
                s.class_name,
                s.score_chinese,
                s.score_math,
                s.score_english,
                s.score_science,
                s.score_society,
                s.total_score,
                s.is_included,
                COALESCE(layer_map.layer_code, 'C') AS layer_code
            FROM biz_scores s
            JOIN biz_students st ON s.student_id = st.id
            LEFT JOIN (
                SELECT
                    cld.class_name,
                    CASE
                        WHEN MAX(CASE WHEN cl.layer_name LIKE 'A%' THEN 1 ELSE 0 END) = 1 THEN 'A'
                        WHEN MAX(CASE WHEN cl.layer_name LIKE 'B%' THEN 1 ELSE 0 END) = 1 THEN 'B'
                        WHEN MAX(CASE WHEN cl.layer_name LIKE 'C%' THEN 1 ELSE 0 END) = 1 THEN 'C'
                        ELSE 'C'
                    END AS layer_code
                FROM biz_class_layer_details cld
                JOIN biz_class_layers cl ON cld.layer_id = cl.id
                WHERE cl.exam_id = :exam_id
                GROUP BY cld.class_name
            ) layer_map ON s.class_name = layer_map.class_name
            WHERE s.exam_id = :exam_id
              AND s.is_included = 1
        """
        score_rows = db.execute(text(scores_sql), {
            "exam_id": request.exam_id,
            "grade_level": request.grade_level
        }).fetchall()
        scores_results = [_row_with_subject_scores(row) for row in score_rows]
        
        if not scores_results:
            return {
                "success": False,
                "message": "该考试暂无有效成绩数据"
            }
        
        # 根据分析类型执行不同的分析算法
        analysis_data = {}
        
        if request.analysis_type == 'overall':
            analysis_data = analyze_overall(scores_results, db)
        elif request.analysis_type == 'layer_comparison':
            analysis_data = analyze_layer_comparison(scores_results, request.analysis_scope, db)
        elif request.analysis_type == 'subject_analysis':
            analysis_data = analyze_subjects(scores_results, db)
        elif request.analysis_type == 'student_progress':
            analysis_data = analyze_student_progress(request.exam_id, request.grade_level, db)
        elif request.analysis_type == 'class_contrast':
            analysis_data = analyze_class_contrast(scores_results, db)
        else:
            raise HTTPException(status_code=400, detail="未知的分析类型")
        
        # 生成分析ID
        analysis_id = f"ANALYSIS_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
        
        # 保存分析结果
        insert_sql = """
            INSERT INTO biz_score_analysis 
            (analysis_id, exam_id, exam_name, grade_level, analysis_type, 
             analysis_scope, analysis_data, created_by, created_by_name, status)
            VALUES 
            (:analysis_id, :exam_id, :exam_name, :grade_level, :analysis_type,
             :analysis_scope, :analysis_data, :created_by, :created_by_name, 'draft')
        """
        
        db.execute(text(insert_sql), {
            "analysis_id": analysis_id,
            "exam_id": request.exam_id,
            "exam_name": exam_name,
            "grade_level": request.grade_level,
            "analysis_type": request.analysis_type,
            "analysis_scope": request.analysis_scope,
            "analysis_data": json.dumps(analysis_data, ensure_ascii=False, default=str),
            "created_by": current_user["id"],
            "created_by_name": current_user.get("real_name", current_user["username"])
        })
        
        db.commit()
        
        # 记录操作日志
        log_sql = """
            INSERT INTO biz_analysis_logs 
            (analysis_id, action_type, action_by, action_by_name, action_by_role, 
             action_detail, ip_address, user_agent)
            VALUES 
            (:analysis_id, 'create_analysis', :action_by, :action_by_name, :action_by_role,
             :action_detail, :ip_address, :user_agent)
        """
        db.execute(text(log_sql), {
            "analysis_id": analysis_id,
            "action_by": current_user["id"],
            "action_by_name": current_user.get("real_name", current_user["username"]),
            "action_by_role": current_user.get("role_name", current_user.get("permission_code", "")),
            "action_detail": json.dumps(request.dict(), ensure_ascii=False),
            "ip_address": req.client.host if req.client else None,
            "user_agent": req.headers.get("user-agent")
        })
        db.commit()
        
        return {
            "success": True,
            "analysis_id": analysis_id,
            "message": "分析完成",
            "data": analysis_data
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"成绩分析失败: {e}")
        raise HTTPException(status_code=500, detail="成绩分析失败")


def analyze_overall(scores_results, db) -> Dict[str, Any]:
    """整体分析"""
    total_scores = [r.total_score for r in scores_results if r.total_score is not None]
    
    # 基础统计
    basic_stats = calculate_basic_statistics(total_scores)
    
    # 成绩分布
    thresholds_sql = "SELECT config_value FROM biz_analysis_configs WHERE config_key = 'score_thresholds'"
    thresholds_result = db.execute(text(thresholds_sql)).fetchone()
    thresholds = json.loads(thresholds_result.config_value) if thresholds_result else {
        "excellent": 90, "good": 80, "pass": 60
    }
    
    distribution = calculate_score_distribution(total_scores, thresholds)
    
    # 班级统计
    class_stats = {}
    for result in scores_results:
        class_name = result.class_name
        if class_name not in class_stats:
            class_stats[class_name] = []
        if result.total_score is not None:
            class_stats[class_name].append(result.total_score)
    
    class_analysis = {}
    for class_name, scores in class_stats.items():
        class_analysis[class_name] = calculate_basic_statistics(scores)
    
    return {
        "basic_statistics": basic_stats,
        "score_distribution": distribution,
        "class_analysis": class_analysis,
        "total_students": len(scores_results),
        "valid_scores": len(total_scores)
    }


def analyze_layer_comparison(scores_results, analysis_scope, db) -> Dict[str, Any]:
    """层次对比分析"""
    # 按层次分组
    layer_groups = {'A': [], 'B': [], 'C': []}
    
    for result in scores_results:
        layer_code = result.layer_code or 'B'  # 默认平行班
        if result.total_score is not None:
            layer_groups[layer_code].append(result.total_score)
    
    # 如果指定了特定层次
    if analysis_scope != 'all':
        target_layer = analysis_scope.replace('layer_', '').upper()
        layer_groups = {k: v for k, v in layer_groups.items() if k == target_layer}
    
    # 计算各层次统计
    layer_stats = {}
    for layer_code, scores in layer_groups.items():
        if scores:
            layer_stats[layer_code] = calculate_basic_statistics(scores)
    
    # 层次间对比
    comparisons = {}
    layer_codes = list(layer_stats.keys())
    for i in range(len(layer_codes)):
        for j in range(i + 1, len(layer_codes)):
            layer1, layer2 = layer_codes[i], layer_codes[j]
            scores1 = layer_groups[layer1]
            scores2 = layer_groups[layer2]
            
            if scores1 and scores2:
                t_test_result = perform_t_test(scores1, scores2)
                comparisons[f"{layer1}_vs_{layer2}"] = t_test_result
    
    return {
        "layer_statistics": layer_stats,
        "layer_comparisons": comparisons,
        "sample_sizes": {k: len(v) for k, v in layer_groups.items()}
    }


def analyze_subjects(scores_results, db) -> Dict[str, Any]:
    """学科分析"""
    # 获取考试科目
    if not scores_results:
        return {}
    
    # 解析成绩JSON
    subject_scores = {}
    for result in scores_results:
        if result.scores:
            scores_dict = json.loads(result.scores) if isinstance(result.scores, str) else result.scores
            for subject, score in scores_dict.items():
                if subject not in subject_scores:
                    subject_scores[subject] = []
                if score is not None:
                    subject_scores[subject].append(float(score))
    
    # 计算各学科统计
    subject_stats = {}
    for subject, scores in subject_scores.items():
        subject_stats[subject] = calculate_basic_statistics(scores)
    
    return {
        "subject_statistics": subject_stats,
        "subject_count": len(subject_stats)
    }


def analyze_student_progress(current_exam_id: int, grade_level: str, db) -> Dict[str, Any]:
    """学生进退步分析"""
    # 获取当前考试和上一次考试
    exam_sql = """
        SELECT id, exam_name, exam_date 
        FROM biz_exams 
        WHERE grade_level = :grade_level
          AND (
            id = :current_exam_id
            OR exam_date < (SELECT exam_date FROM biz_exams WHERE id = :current_exam_id)
          )
        ORDER BY exam_date DESC
        LIMIT 2
    """
    exam_results = db.execute(
        text(exam_sql),
        {"grade_level": grade_level, "current_exam_id": current_exam_id}
    ).fetchall()
    
    if len(exam_results) < 2:
        return {"message": "历史考试数据不足，无法分析进退步"}
    
    current_exam = exam_results[0]
    previous_exam = exam_results[1]
    
    # 获取两次考试的成绩
    scores_sql = """
        SELECT s1.student_id, st.name as student_name, s1.total_score as current_score,
               s2.total_score as previous_score
        FROM biz_scores s1
        JOIN biz_students st ON s1.student_id = st.id
        JOIN biz_scores s2 ON s1.student_id = s2.student_id
        WHERE s1.exam_id = :current_exam_id AND s2.exam_id = :previous_exam_id
          AND s1.is_included = 1 AND s2.is_included = 1
    """
    scores_results = db.execute(text(scores_sql), {
        "current_exam_id": current_exam.id,
        "previous_exam_id": previous_exam.id
    }).fetchall()
    
    # 计算进退步
    progress_data = []
    for result in scores_results:
        if result.current_score is not None and result.previous_score is not None:
            score_change = result.current_score - result.previous_score
            progress_data.append({
                "student_id": result.student_id,
                "student_name": result.student_name,
                "current_score": result.current_score,
                "previous_score": result.previous_score,
                "score_change": score_change,
                "change_percentage": (score_change / result.previous_score * 100) if result.previous_score > 0 else 0
            })
    
    # 排序
    progress_data.sort(key=lambda x: x["score_change"], reverse=True)
    
    # 统计
    improved = len([x for x in progress_data if x["score_change"] > 0])
    declined = len([x for x in progress_data if x["score_change"] < 0])
    unchanged = len([x for x in progress_data if x["score_change"] == 0])
    
    return {
        "current_exam": current_exam.exam_name,
        "previous_exam": previous_exam.exam_name,
        "total_students": len(progress_data),
        "improved_count": improved,
        "declined_count": declined,
        "unchanged_count": unchanged,
        "top_improved": progress_data[:10],  # 进步最大的10人
        "top_declined": progress_data[-10:]   # 退步最大的10人
    }


def analyze_class_contrast(scores_results, db) -> Dict[str, Any]:
    """班级对比分析"""
    # 按班级分组
    class_groups = {}
    for result in scores_results:
        class_name = result.class_name
        if class_name not in class_groups:
            class_groups[class_name] = []
        if result.total_score is not None:
            class_groups[class_name].append(result.total_score)
    
    # 计算各班统计
    class_stats = {}
    for class_name, scores in class_groups.items():
        class_stats[class_name] = calculate_basic_statistics(scores)
    
    # 计算年级平均
    all_scores = [score for scores in class_groups.values() for score in scores]
    grade_mean = sum(all_scores) / len(all_scores) if all_scores else 0
    if all_scores:
        variance = sum((score - grade_mean) ** 2 for score in all_scores) / len(all_scores)
        grade_std = math.sqrt(variance)
    else:
        grade_std = 1
    
    # 计算各班Z值
    class_z_scores = {}
    for class_name, scores in class_groups.items():
        if scores:
            class_mean = sum(scores) / len(scores)
            class_z_scores[class_name] = (class_mean - grade_mean) / grade_std if grade_std > 0 else 0
    
    # 排名
    class_ranking = sorted(class_z_scores.items(), key=lambda x: x[1], reverse=True)
    
    return {
        "class_statistics": class_stats,
        "class_z_scores": class_z_scores,
        "class_ranking": [{"class_name": name, "z_score": score} for name, score in class_ranking],
        "grade_mean": float(grade_mean),
        "grade_std": float(grade_std)
    }


def _excel_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def _safe_sheet_title(title: str, used_titles: Optional[set] = None) -> str:
    used_titles = used_titles if used_titles is not None else set()
    cleaned = "".join("_" if char in "[]:*?/\\" else char for char in str(title or "Sheet")).strip()
    cleaned = cleaned[:31] or "Sheet"
    candidate = cleaned
    index = 2
    while candidate in used_titles:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(candidate)
    return candidate


def _append_key_value_rows(sheet, data: Dict[str, Any]):
    sheet.append(["指标", "数值"])
    for key, value in data.items():
        sheet.append([key, _excel_cell_value(value)])


def _append_list_rows(sheet, rows: List[Any]):
    dict_rows = [row for row in rows if isinstance(row, dict)]
    if dict_rows and len(dict_rows) == len(rows):
        headers = []
        for row in dict_rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)
        sheet.append(headers)
        for row in dict_rows:
            sheet.append([_excel_cell_value(row.get(header)) for header in headers])
        return

    sheet.append(["序号", "内容"])
    for index, row in enumerate(rows, start=1):
        sheet.append([index, _excel_cell_value(row)])


def _append_nested_mapping_rows(sheet, data: Dict[str, Any]):
    child_dicts = {key: value for key, value in data.items() if isinstance(value, dict)}
    if child_dicts and len(child_dicts) == len(data):
        headers = ["项目"]
        for child in child_dicts.values():
            for key in child.keys():
                if key not in headers:
                    headers.append(key)
        sheet.append(headers)
        for item, child in child_dicts.items():
            sheet.append([item, *[_excel_cell_value(child.get(header)) for header in headers[1:]]])
        return

    _append_key_value_rows(sheet, data)


def _append_export_sheet(sheet, value: Any):
    if isinstance(value, dict):
        _append_nested_mapping_rows(sheet, value)
    elif isinstance(value, list):
        _append_list_rows(sheet, value)
    else:
        sheet.append(["内容"])
        sheet.append([_excel_cell_value(value)])


def _build_analysis_workbook(result, analysis_data: Dict[str, Any]) -> bytes:
    workbook = Workbook()
    used_titles = set()
    summary = workbook.active
    summary.title = _safe_sheet_title("分析概览", used_titles)
    summary.append(["字段", "值"])
    summary.append(["分析ID", result.analysis_id])
    summary.append(["考试", result.exam_name])
    summary.append(["年级", result.grade_level])
    summary.append(["分析类型", result.analysis_type])
    summary.append(["分析范围", result.analysis_scope])
    summary.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    for key, value in (analysis_data or {}).items():
        if not isinstance(value, (dict, list)):
            summary.append([key, _excel_cell_value(value)])

    for key, value in (analysis_data or {}).items():
        if isinstance(value, (dict, list)):
            sheet = workbook.create_sheet(_safe_sheet_title(key, used_titles))
            _append_export_sheet(sheet, value)

    raw_sheet = workbook.create_sheet(_safe_sheet_title("原始JSON", used_titles))
    raw_sheet.append(["analysis_data"])
    raw_sheet.append([json.dumps(analysis_data or {}, ensure_ascii=False, indent=2, default=str)])

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _ensure_reportlab_available() -> None:
    if SimpleDocTemplate is None:
        raise HTTPException(status_code=500, detail="PDF导出依赖未安装，请安装 reportlab")


def _pdf_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    return json.dumps(value, ensure_ascii=False, default=str)


def _flatten_pdf_rows(data: Dict[str, Any], limit: int = 36) -> List[List[str]]:
    rows: List[List[str]] = []
    for key, value in (data or {}).items():
        if len(rows) >= limit:
            break
        if isinstance(value, dict):
            rows.append([str(key), f"{len(value)} 项"])
        elif isinstance(value, list):
            rows.append([str(key), f"{len(value)} 条"])
        else:
            rows.append([str(key), _pdf_text(value)])
    return rows or [["内容", "暂无数据"]]


def _build_analysis_pdf(result, analysis_data: Dict[str, Any]) -> bytes:
    _ensure_reportlab_available()
    stream = io.BytesIO()
    if UnicodeCIDFont is not None:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
            font_name = "STSong-Light"
        except Exception:
            font_name = "Helvetica"
    else:
        font_name = "Helvetica"

    doc = SimpleDocTemplate(
        stream,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = font_name
    story = [
        Paragraph(f"{_pdf_text(result.exam_name)} 成绩分析报告", styles["Title"]),
        Spacer(1, 12),
        Paragraph(f"年级：{_pdf_text(result.grade_level)}　范围：{_pdf_text(getattr(result, 'analysis_scope', 'full'))}", styles["Normal"]),
        Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]),
        Spacer(1, 16),
    ]

    summary_table = Table(
        [["指标", "结果"], *_flatten_pdf_rows(analysis_data)],
        colWidths=[150, 330],
        hAlign="LEFT",
    )
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 16))

    for key, value in (analysis_data or {}).items():
        if not isinstance(value, dict):
            continue
        child_rows = _flatten_pdf_rows(value, limit=12)
        story.append(Paragraph(str(key), styles["Heading3"]))
        table = Table([["项目", "结果"], *child_rows], colWidths=[150, 330], hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

    doc.build(story)
    return stream.getvalue()


def _ensure_score_analysis_bundle_table(db: Session) -> None:
    if is_postgresql(db) or is_sqlite(db):
        bundle_id_type = "BIGSERIAL" if is_postgresql(db) else "INTEGER"
        db.execute(text(f"""
            CREATE TABLE IF NOT EXISTS biz_score_analysis_bundles (
              id {bundle_id_type} PRIMARY KEY,
              bundle_id VARCHAR(80) NOT NULL UNIQUE,
              exam_id BIGINT NOT NULL,
              exam_name VARCHAR(120) NOT NULL,
              grade_level VARCHAR(20) NOT NULL,
              status VARCHAR(20) DEFAULT 'ready',
              result_json TEXT NOT NULL,
              source_hash VARCHAR(80) DEFAULT NULL,
              generated_by BIGINT DEFAULT NULL,
              generated_by_name VARCHAR(80) DEFAULT NULL,
              generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
              updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
              CONSTRAINT uk_exam_grade_bundle UNIQUE (exam_id, grade_level)
            )
        """))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_grade_generated_at ON biz_score_analysis_bundles (grade_level, generated_at)"))
    else:
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS biz_score_analysis_bundles (
              id BIGINT AUTO_INCREMENT PRIMARY KEY,
              bundle_id VARCHAR(80) NOT NULL UNIQUE COMMENT '结果包ID',
              exam_id BIGINT NOT NULL COMMENT '考试ID',
              exam_name VARCHAR(120) NOT NULL COMMENT '考试名称',
              grade_level VARCHAR(20) NOT NULL COMMENT '年级',
              status VARCHAR(20) DEFAULT 'ready' COMMENT '结果包状态',
              result_json LONGTEXT NOT NULL COMMENT '结果包JSON',
              source_hash VARCHAR(80) DEFAULT NULL COMMENT '来源数据版本',
              generated_by BIGINT DEFAULT NULL COMMENT '生成人',
              generated_by_name VARCHAR(80) DEFAULT NULL COMMENT '生成人姓名',
              generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
              updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              UNIQUE KEY uk_exam_grade_bundle (exam_id, grade_level),
              INDEX idx_grade_generated_at (grade_level, generated_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='成绩分析结果包缓存表'
        """))
    db.commit()


def _parse_subjects(value: Any) -> List[str]:
    if not value:
        return []
    try:
        parsed = json.loads(value) if isinstance(value, str) else value
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def _fetch_bundle_exam(db: Session, exam_id: int) -> Any:
    exam = db.execute(text("""
        SELECT id, exam_name, grade_level, subjects, full_score, exam_date, term
        FROM biz_exams
        WHERE id = :exam_id
    """), {"exam_id": exam_id}).fetchone()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    return exam


def _fetch_bundle_scores(db: Session, exam_id: int) -> List[Any]:
    rows = db.execute(text("""
        SELECT
            s.id,
            s.exam_id,
            s.student_id,
            st.name AS student_name,
            s.exam_number,
            s.class_name,
            s.score_chinese,
            s.score_math,
            s.score_english,
            s.score_science,
            s.score_society,
            s.total_score,
            s.is_included,
            COALESCE(layer_map.layer_code, 'C') AS layer_code
        FROM biz_scores s
        LEFT JOIN biz_students st ON s.student_id = st.id
        LEFT JOIN (
            SELECT
                cld.class_name,
                CASE
                    WHEN MAX(CASE WHEN cl.layer_name LIKE '%A%' THEN 1 ELSE 0 END) = 1 THEN 'A'
                    WHEN MAX(CASE WHEN cl.layer_name LIKE '%B%' THEN 1 ELSE 0 END) = 1 THEN 'B'
                    WHEN MAX(CASE WHEN cl.layer_name LIKE '%C%' THEN 1 ELSE 0 END) = 1 THEN 'C'
                    ELSE 'C'
                END AS layer_code
            FROM biz_class_layer_details cld
            JOIN biz_class_layers cl ON cld.layer_id = cl.id
            WHERE cl.exam_id = :exam_id
            GROUP BY cld.class_name
        ) layer_map ON s.class_name = layer_map.class_name
        WHERE s.exam_id = :exam_id
          AND s.is_included = 1
        ORDER BY s.total_score DESC
    """), {"exam_id": exam_id}).fetchall()
    return [_row_with_subject_scores(row) for row in rows]


def _build_result_bundle_data(exam: Any, scores_results: List[Any], db: Session) -> Dict[str, Any]:
    if not scores_results:
        raise HTTPException(status_code=400, detail="该考试暂无有效成绩数据")

    progress_data: Dict[str, Any]
    try:
        progress_data = analyze_student_progress(exam.id, exam.grade_level, db)
    except Exception as exc:
        progress_data = {"message": f"历史数据暂不可用：{exc}"}

    return {
        "exam": {
            "id": exam.id,
            "exam_name": exam.exam_name,
            "grade_level": exam.grade_level,
            "term": getattr(exam, "term", ""),
            "exam_date": exam.exam_date.isoformat() if exam.exam_date else "",
            "subjects": _parse_subjects(exam.subjects),
            "full_score": float(exam.full_score) if exam.full_score is not None else None,
        },
        "modules": {
            "overall": analyze_overall(scores_results, db),
            "layer_comparison": analyze_layer_comparison(scores_results, "all", db),
            "subject_analysis": analyze_subjects(scores_results, db),
            "class_contrast": analyze_class_contrast(scores_results, db),
            "student_progress": progress_data,
        },
        "generated_at": datetime.now().isoformat(),
    }


def _store_result_bundle(
    db: Session,
    exam: Any,
    result_json: Dict[str, Any],
    current_user: Dict[str, Any],
) -> str:
    _ensure_score_analysis_bundle_table(db)
    bundle_id = f"BUNDLE_{exam.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:6]}"
    source_hash = str(uuid.uuid5(uuid.NAMESPACE_DNS, json.dumps({
        "exam_id": exam.id,
        "generated_at": result_json.get("generated_at"),
    }, ensure_ascii=False)))
    if is_postgresql(db) or is_sqlite(db):
        store_sql = """
            INSERT INTO biz_score_analysis_bundles
              (bundle_id, exam_id, exam_name, grade_level, status, result_json,
               source_hash, generated_by, generated_by_name, generated_at, updated_at)
            VALUES
              (:bundle_id, :exam_id, :exam_name, :grade_level, 'ready', :result_json,
               :source_hash, :generated_by, :generated_by_name, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT (exam_id, grade_level) DO UPDATE SET
              bundle_id = excluded.bundle_id,
              exam_name = excluded.exam_name,
              status = 'ready',
              result_json = excluded.result_json,
              source_hash = excluded.source_hash,
              generated_by = excluded.generated_by,
              generated_by_name = excluded.generated_by_name,
              generated_at = CURRENT_TIMESTAMP,
              updated_at = CURRENT_TIMESTAMP
        """
    else:
        store_sql = """
            INSERT INTO biz_score_analysis_bundles
              (bundle_id, exam_id, exam_name, grade_level, status, result_json,
               source_hash, generated_by, generated_by_name, generated_at)
            VALUES
              (:bundle_id, :exam_id, :exam_name, :grade_level, 'ready', :result_json,
               :source_hash, :generated_by, :generated_by_name, NOW())
            ON DUPLICATE KEY UPDATE
              bundle_id = VALUES(bundle_id),
              exam_name = VALUES(exam_name),
              status = 'ready',
              result_json = VALUES(result_json),
              source_hash = VALUES(source_hash),
              generated_by = VALUES(generated_by),
              generated_by_name = VALUES(generated_by_name),
              generated_at = NOW()
        """
    db.execute(text(store_sql), {
        "bundle_id": bundle_id,
        "exam_id": exam.id,
        "exam_name": exam.exam_name,
        "grade_level": exam.grade_level,
        "result_json": json.dumps(result_json, ensure_ascii=False, default=str),
        "source_hash": source_hash,
        "generated_by": current_user.get("id"),
        "generated_by_name": current_user.get("real_name", current_user.get("username", "")),
    })
    db.commit()
    return bundle_id


def _bundle_row_to_payload(row: Any, current_user: Dict[str, Any], db: Session) -> Dict[str, Any]:
    settings = fetch_score_visibility_settings(db)
    visibility = resolve_visibility_for_user(current_user, settings)
    result_json = json.loads(row.result_json or "{}")
    filtered_result = filter_rank_fields_by_visibility(result_json, visibility)
    return {
        "bundle_id": row.bundle_id,
        "exam_id": row.exam_id,
        "exam_name": row.exam_name,
        "grade_level": row.grade_level,
        "status": row.status,
        "generated_at": row.generated_at.isoformat() if row.generated_at else "",
        "updated_at": row.updated_at.isoformat() if row.updated_at else "",
        "visibility": visibility,
        "result": filtered_result,
    }


# ============ 结果包查询API ============

@router.post("/bundles/{exam_id}/refresh")
def refresh_analysis_bundle(
    exam_id: int,
    request: BundleRefreshRequest,
    current_user: dict = Depends(require_permission_codes(*SCORE_ANALYSIS_EXECUTE_PERMISSION_CODES)),
    db: Session = Depends(get_db),
):
    """一键刷新指定考试的成绩分析结果包。"""
    try:
        exam = _fetch_bundle_exam(db, exam_id)
        grade_level = request.grade_level or exam.grade_level
        accessible_grades = get_accessible_grades(current_user, db)
        if grade_level not in accessible_grades:
            raise HTTPException(status_code=403, detail="无权刷新该年级结果")
        if exam.grade_level and exam.grade_level != grade_level:
            raise HTTPException(status_code=400, detail="考试年级与刷新年级不一致")

        scores_results = _fetch_bundle_scores(db, exam_id)
        result_json = _build_result_bundle_data(exam, scores_results, db)
        bundle_id = _store_result_bundle(db, exam, result_json, current_user)
        row = db.execute(text("""
            SELECT *
            FROM biz_score_analysis_bundles
            WHERE bundle_id = :bundle_id
        """), {"bundle_id": bundle_id}).fetchone()

        return {
            "success": True,
            "message": "成绩分析结果包已刷新",
            "data": _bundle_row_to_payload(row, current_user, db),
        }
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        logger.error(f"刷新成绩分析结果包失败: {exc}")
        raise HTTPException(status_code=500, detail="刷新成绩分析结果包失败")


@router.get("/bundles/latest")
def get_latest_analysis_bundle(
    grade_level: Optional[str] = Query(None),
    exam_id: Optional[int] = Query(None),
    scope: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """查询最新成绩分析结果包，返回前会按当前角色过滤排名字段。"""
    try:
        _ensure_score_analysis_bundle_table(db)
        params: Dict[str, Any] = {}
        conditions: List[str] = ["status = 'ready'"]

        if exam_id:
            conditions.append("exam_id = :exam_id")
            params["exam_id"] = exam_id
        if grade_level:
            conditions.append("grade_level = :grade_level")
            params["grade_level"] = grade_level

        accessible_grades = get_accessible_grades(current_user, db)
        if accessible_grades and grade_level and grade_level not in accessible_grades:
            raise HTTPException(status_code=403, detail="无权访问该年级结果")
        if accessible_grades and not grade_level and current_user.get("permission_code") not in {"sys_admin", "edu_admin"}:
            _add_in_condition(conditions, params, "grade_level", accessible_grades, "bundle_grade")
        if not accessible_grades and current_user.get("permission_code") not in {"sys_admin", "edu_admin"}:
            return {
                "success": False,
                "message": "当前角色暂无可查询的年级结果包",
                "data": None,
            }

        row = db.execute(text(f"""
            SELECT *
            FROM biz_score_analysis_bundles
            WHERE {" AND ".join(conditions)}
            ORDER BY generated_at DESC
            LIMIT 1
        """), params).fetchone()

        if not row:
            return {
                "success": False,
                "message": "暂无已生成的成绩分析结果包，请先点击更新结果",
                "data": None,
            }

        payload = _bundle_row_to_payload(row, current_user, db)
        if scope:
            payload["scope"] = scope
        return {
            "success": True,
            "data": payload,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"查询成绩分析结果包失败: {exc}")
        raise HTTPException(status_code=500, detail="查询成绩分析结果包失败")


@router.get("/bundles/{bundle_id}/export")
def export_analysis_bundle(
    bundle_id: str,
    format: str = Query("excel", regex="^(excel|pdf|json)$"),
    view: str = Query("full"),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出已缓存的成绩分析结果包。"""
    try:
        _ensure_score_analysis_bundle_table(db)
        row = db.execute(text("""
            SELECT *
            FROM biz_score_analysis_bundles
            WHERE bundle_id = :bundle_id
        """), {"bundle_id": bundle_id}).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="结果包不存在")

        settings = fetch_score_visibility_settings(db)
        visibility = resolve_visibility_for_user(current_user, settings)
        if not visibility.get("allow_export") and current_user.get("permission_code") not in {"sys_admin", "edu_admin"}:
            raise HTTPException(status_code=403, detail="当前角色未开放成绩报告导出")

        payload = _bundle_row_to_payload(row, current_user, db)
        analysis_data = payload["result"]
        result = SimpleNamespace(
            analysis_id=bundle_id,
            exam_name=row.exam_name,
            grade_level=row.grade_level,
            analysis_type=f"bundle:{view}",
            analysis_scope=view,
        )

        if format == "json":
            return {"success": True, "data": analysis_data}

        if format == "excel":
            excel_data = _build_analysis_workbook(result, analysis_data)
            return StreamingResponse(
                io.BytesIO(excel_data),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=score_analysis_bundle_{bundle_id}.xlsx"},
            )

        pdf_data = _build_analysis_pdf(result, analysis_data)
        return StreamingResponse(
            io.BytesIO(pdf_data),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=score_analysis_bundle_{bundle_id}.pdf"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"导出成绩分析结果包失败: {exc}")
        raise HTTPException(status_code=500, detail="导出成绩分析结果包失败")


# ============ 分析结果查询API ============

@router.get("/results")
def get_analysis_results(
    exam_id: Optional[int] = Query(None),
    grade_level: Optional[str] = Query(None),
    analysis_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(require_permission_codes(*SCORE_ANALYSIS_VIEW_PERMISSION_CODES)),
    db: Session = Depends(get_db)
):
    """获取分析结果列表"""
    try:
        # 检查权限
        accessible_grades = get_accessible_grades(current_user, db)
        if grade_level and grade_level not in accessible_grades:
            raise HTTPException(status_code=403, detail="无权访问该年级数据")
        
        conditions = []
        params = {}
        
        if exam_id:
            conditions.append("exam_id = :exam_id")
            params["exam_id"] = exam_id
        
        if grade_level:
            conditions.append("grade_level = :grade_level")
            params["grade_level"] = grade_level
        else:
            if not _add_in_condition(conditions, params, "grade_level", accessible_grades, "grade"):
                return {
                    "success": True,
                    "total": 0,
                    "page": page,
                    "page_size": page_size,
                    "records": []
                }
        
        if analysis_type:
            conditions.append("analysis_type = :analysis_type")
            params["analysis_type"] = analysis_type
        
        if status:
            conditions.append("status = :status")
            params["status"] = status
        
        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""
        
        # 查询总数
        count_sql = f"SELECT COUNT(*) as total FROM biz_score_analysis {where_clause}"
        count_result = db.execute(text(count_sql), params).fetchone()
        total = count_result.total if count_result else 0
        
        # 查询列表
        offset = (page - 1) * page_size
        list_sql = f"""
            SELECT * FROM biz_score_analysis 
            {where_clause}
            ORDER BY created_at DESC
            LIMIT :limit OFFSET :offset
        """
        params["limit"] = page_size
        params["offset"] = offset
        
        results = db.execute(text(list_sql), params).fetchall()
        
        records = []
        for result in results:
            records.append({
                "analysis_id": result.analysis_id,
                "exam_id": result.exam_id,
                "exam_name": result.exam_name,
                "grade_level": result.grade_level,
                "analysis_type": result.analysis_type,
                "analysis_scope": result.analysis_scope,
                "created_by": result.created_by,
                "created_by_name": result.created_by_name,
                "created_at": result.created_at.isoformat() if result.created_at else "",
                "status": result.status
            })
        
        return {
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "records": records
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取分析结果列表失败: {e}")
        raise HTTPException(status_code=500, detail="获取分析结果列表失败")


@router.get("/results/{analysis_id}")
def get_analysis_detail(
    analysis_id: str,
    current_user: dict = Depends(require_permission_codes(*SCORE_ANALYSIS_VIEW_PERMISSION_CODES)),
    db: Session = Depends(get_db)
):
    """获取分析结果详情"""
    try:
        sql = "SELECT * FROM biz_score_analysis WHERE analysis_id = :analysis_id"
        result = db.execute(text(sql), {"analysis_id": analysis_id}).fetchone()
        
        if not result:
            raise HTTPException(status_code=404, detail="分析结果不存在")
        
        # 检查权限
        accessible_grades = get_accessible_grades(current_user, db)
        if result.grade_level not in accessible_grades:
            raise HTTPException(status_code=403, detail="无权访问该分析结果")
        
        return {
            "success": True,
            "data": {
                "analysis_id": result.analysis_id,
                "exam_id": result.exam_id,
                "exam_name": result.exam_name,
                "grade_level": result.grade_level,
                "analysis_type": result.analysis_type,
                "analysis_scope": result.analysis_scope,
                "analysis_data": json.loads(result.analysis_data) if result.analysis_data else {},
                "created_by": result.created_by,
                "created_by_name": result.created_by_name,
                "created_at": result.created_at.isoformat() if result.created_at else "",
                "status": result.status
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取分析结果详情失败: {e}")
        raise HTTPException(status_code=500, detail="获取分析结果详情失败")


# ============ 成果发布API ============

@router.post("/publish")
def publish_analysis(
    request: PublicationRequest,
    req: Request,
    current_user: dict = Depends(require_permission("analysis_publish")),
    db: Session = Depends(get_db)
):
    """发布分析成果"""
    try:
        # 获取分析结果
        analysis_sql = "SELECT * FROM biz_score_analysis WHERE analysis_id = :analysis_id"
        analysis_result = db.execute(text(analysis_sql), {
            "analysis_id": request.analysis_id
        }).fetchone()
        
        if not analysis_result:
            raise HTTPException(status_code=404, detail="分析结果不存在")

        accessible_grades = get_accessible_grades(current_user, db)
        if analysis_result.grade_level not in accessible_grades:
            raise HTTPException(status_code=403, detail="无权发布该年级分析结果")
        
        # 生成发布ID
        publication_id = f"PUB_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
        
        # 获取接收用户列表
        recipient_types = request.recipient_types
        recipient_permission_codes = []
        for recipient_type in recipient_types:
            recipient_permission_codes.extend(
                RECIPIENT_TYPE_PERMISSION_CODES.get(recipient_type, (recipient_type,))
            )
        recipient_permission_codes = sorted(set(recipient_permission_codes))

        if not recipient_permission_codes:
            raise HTTPException(status_code=400, detail="请选择有效接收对象")

        placeholders = ', '.join([f':type_{i}' for i in range(len(recipient_permission_codes))])
        type_params = {f'type_{i}': t for i, t in enumerate(recipient_permission_codes)}
        
        users_sql = f"""
            SELECT u.id, u.username, u.real_name, r.role_name, r.permission_code
            FROM sys_users u
            JOIN sys_roles r ON u.role_id = r.id
            WHERE r.permission_code IN ({placeholders})
              AND u.is_active = 1
        """
        users_results = db.execute(text(users_sql), type_params).fetchall()
        
        # 保存发布记录
        insert_pub_sql = """
            INSERT INTO biz_analysis_publications 
            (publication_id, analysis_id, exam_id, exam_name, grade_level,
             title, content_summary, published_by, published_by_name,
             recipient_types, recipient_count, status)
            VALUES 
            (:publication_id, :analysis_id, :exam_id, :exam_name, :grade_level,
             :title, :content_summary, :published_by, :published_by_name,
             :recipient_types, :recipient_count, 'active')
        """
        
        db.execute(text(insert_pub_sql), {
            "publication_id": publication_id,
            "analysis_id": request.analysis_id,
            "exam_id": analysis_result.exam_id,
            "exam_name": analysis_result.exam_name,
            "grade_level": analysis_result.grade_level,
            "title": request.title,
            "content_summary": request.content_summary,
            "published_by": current_user["id"],
            "published_by_name": current_user.get("real_name", current_user["username"]),
            "recipient_types": json.dumps(recipient_types, ensure_ascii=False),
            "recipient_count": len(users_results)
        })
        
        # 保存接收对象明细
        for user in users_results:
            insert_recipient_sql = """
                INSERT INTO biz_publication_recipients 
                (publication_id, user_id, user_name, user_role)
                VALUES 
                (:publication_id, :user_id, :user_name, :user_role)
            """
            db.execute(text(insert_recipient_sql), {
                "publication_id": publication_id,
                "user_id": user.id,
                "user_name": user.real_name or user.username,
                "user_role": user.permission_code
            })
        
        # 更新分析结果状态为已发布
        update_sql = "UPDATE biz_score_analysis SET status = 'published' WHERE analysis_id = :analysis_id"
        db.execute(text(update_sql), {"analysis_id": request.analysis_id})
        
        db.commit()
        
        # 记录日志
        log_sql = """
            INSERT INTO biz_analysis_logs 
            (analysis_id, publication_id, action_type, action_by, action_by_name,
             action_by_role, action_detail, ip_address, user_agent)
            VALUES 
            (:analysis_id, :publication_id, 'publish', :action_by, :action_by_name,
             :action_by_role, :action_detail, :ip_address, :user_agent)
        """
        db.execute(text(log_sql), {
            "analysis_id": request.analysis_id,
            "publication_id": publication_id,
            "action_by": current_user["id"],
            "action_by_name": current_user.get("real_name", current_user["username"]),
            "action_by_role": current_user.get("role_name", current_user.get("permission_code", "")),
            "action_detail": json.dumps(request.dict(), ensure_ascii=False),
            "ip_address": req.client.host if req.client else None,
            "user_agent": req.headers.get("user-agent")
        })
        db.commit()
        
        return {
            "success": True,
            "publication_id": publication_id,
            "message": f"发布成功，共发送给 {len(users_results)} 位接收人"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"发布分析成果失败: {e}")
        raise HTTPException(status_code=500, detail="发布分析成果失败")


@router.get("/publications")
def get_publications(
    exam_id: Optional[int] = Query(None),
    grade_level: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(require_permission_codes(*SCORE_ANALYSIS_VIEW_PERMISSION_CODES)),
    db: Session = Depends(get_db)
):
    """获取发布记录列表"""
    try:
        conditions = ["status = 'active'"]
        params = {}
        accessible_grades = get_accessible_grades(current_user, db)
        
        if exam_id:
            conditions.append("exam_id = :exam_id")
            params["exam_id"] = exam_id
        
        if grade_level:
            if grade_level not in accessible_grades:
                raise HTTPException(status_code=403, detail="无权访问该年级发布记录")
            conditions.append("grade_level = :grade_level")
            params["grade_level"] = grade_level
        else:
            if not _add_in_condition(conditions, params, "grade_level", accessible_grades, "grade"):
                return {
                    "success": True,
                    "total": 0,
                    "page": page,
                    "page_size": page_size,
                    "records": []
                }
        
        where_clause = "WHERE " + " AND ".join(conditions)
        
        # 查询总数
        count_sql = f"SELECT COUNT(*) as total FROM biz_analysis_publications {where_clause}"
        count_result = db.execute(text(count_sql), params).fetchone()
        total = count_result.total if count_result else 0
        
        # 查询列表
        offset = (page - 1) * page_size
        list_sql = f"""
            SELECT * FROM biz_analysis_publications 
            {where_clause}
            ORDER BY published_at DESC
            LIMIT :limit OFFSET :offset
        """
        params["limit"] = page_size
        params["offset"] = offset
        
        results = db.execute(text(list_sql), params).fetchall()
        
        records = []
        for result in results:
            records.append({
                "publication_id": result.publication_id,
                "analysis_id": result.analysis_id,
                "exam_id": result.exam_id,
                "exam_name": result.exam_name,
                "grade_level": result.grade_level,
                "title": result.title,
                "published_by": result.published_by,
                "published_by_name": result.published_by_name,
                "published_at": result.published_at.isoformat() if result.published_at else "",
                "recipient_count": result.recipient_count
            })
        
        return {
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "records": records
        }
        
    except Exception as e:
        logger.error(f"获取发布记录失败: {e}")
        raise HTTPException(status_code=500, detail="获取发布记录失败")


# ============ 数据导出API ============

@router.get("/export/{analysis_id}")
def export_analysis(
    analysis_id: str,
    req: Request,
    format: str = Query("excel", regex="^(excel|pdf|json)$"),
    current_user: dict = Depends(require_permission_codes(*SCORE_ANALYSIS_VIEW_PERMISSION_CODES)),
    db: Session = Depends(get_db)
):
    """导出分析结果"""
    try:
        # 获取分析结果
        sql = "SELECT * FROM biz_score_analysis WHERE analysis_id = :analysis_id"
        result = db.execute(text(sql), {"analysis_id": analysis_id}).fetchone()
        
        if not result:
            raise HTTPException(status_code=404, detail="分析结果不存在")
        
        # 检查权限
        accessible_grades = get_accessible_grades(current_user, db)
        if result.grade_level not in accessible_grades:
            raise HTTPException(status_code=403, detail="无权导出该分析结果")
        
        # 记录日志
        log_sql = """
            INSERT INTO biz_analysis_logs 
            (analysis_id, action_type, action_by, action_by_name,
             action_by_role, action_detail, ip_address, user_agent)
            VALUES 
            (:analysis_id, 'download', :action_by, :action_by_name,
             :action_by_role, :action_detail, :ip_address, :user_agent)
        """
        db.execute(text(log_sql), {
            "analysis_id": analysis_id,
            "action_by": current_user["id"],
            "action_by_name": current_user.get("real_name", current_user["username"]),
            "action_by_role": current_user.get("role_name", current_user.get("permission_code", "")),
            "action_detail": json.dumps({"format": format}, ensure_ascii=False),
            "ip_address": req.client.host if req.client else None,
            "user_agent": req.headers.get("user-agent")
        })
        db.commit()
        
        analysis_data = json.loads(result.analysis_data) if result.analysis_data else {}
        
        if format == "json":
            return {
                "success": True,
                "data": analysis_data
            }

        if format == "excel":
            excel_data = _build_analysis_workbook(result, analysis_data)
            return StreamingResponse(
                io.BytesIO(excel_data),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=score_analysis_{analysis_id}.xlsx"
                }
            )

        if format == "pdf":
            pdf_data = _build_analysis_pdf(result, analysis_data)
            return StreamingResponse(
                io.BytesIO(pdf_data),
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename=score_analysis_{analysis_id}.pdf"
                }
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出分析结果失败: {e}")
        raise HTTPException(status_code=500, detail="导出分析结果失败")


# ============ 操作日志API ============

@router.get("/logs")
def get_analysis_logs(
    analysis_id: Optional[str] = Query(None),
    action_type: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(require_permission("analysis_admin")),
    db: Session = Depends(get_db)
):
    """获取操作日志（仅管理员）"""
    try:
        conditions = []
        params = {}
        
        if analysis_id:
            conditions.append("analysis_id = :analysis_id")
            params["analysis_id"] = analysis_id
        
        if action_type:
            conditions.append("action_type = :action_type")
            params["action_type"] = action_type
        
        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""
        
        # 查询总数
        count_sql = f"SELECT COUNT(*) as total FROM biz_analysis_logs {where_clause}"
        count_result = db.execute(text(count_sql), params).fetchone()
        total = count_result.total if count_result else 0
        
        # 查询列表
        offset = (page - 1) * page_size
        list_sql = f"""
            SELECT * FROM biz_analysis_logs 
            {where_clause}
            ORDER BY created_at DESC
            LIMIT :limit OFFSET :offset
        """
        params["limit"] = page_size
        params["offset"] = offset
        
        results = db.execute(text(list_sql), params).fetchall()
        
        records = []
        for result in results:
            records.append({
                "id": result.id,
                "analysis_id": result.analysis_id,
                "publication_id": result.publication_id,
                "action_type": result.action_type,
                "action_by": result.action_by,
                "action_by_name": result.action_by_name,
                "action_by_role": result.action_by_role,
                "action_detail": json.loads(result.action_detail) if result.action_detail else {},
                "ip_address": result.ip_address,
                "created_at": result.created_at.isoformat() if result.created_at else ""
            })
        
        return {
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "records": records
        }
        
    except Exception as e:
        logger.error(f"获取操作日志失败: {e}")
        raise HTTPException(status_code=500, detail="获取操作日志失败")
