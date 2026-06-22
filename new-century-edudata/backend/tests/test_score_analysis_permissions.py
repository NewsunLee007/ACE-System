from types import SimpleNamespace
from io import BytesIO

from openpyxl import load_workbook

from routers.score_analysis_api import (
    _add_in_condition,
    _build_analysis_workbook,
    analyze_class_contrast,
    check_analysis_permission,
    get_accessible_grades,
)


def user(permission_code, user_id=1, **extra):
    return {"id": user_id, "permission_code": permission_code, **extra}


def test_score_analysis_permission_uses_current_permission_code():
    assert check_analysis_permission(user("subject_leader"), ["research_leader"]) is True
    assert check_analysis_permission(user("exam_admin"), ["exam_admin"]) is True
    assert check_analysis_permission(user("teacher"), ["exam_admin"]) is False


def test_accessible_grades_use_permission_code_and_configured_grade():
    assert get_accessible_grades(user("exam_admin"), db=None) == ["7年级", "8年级", "9年级"]
    assert get_accessible_grades(
        user("grade_leader", grade_level="8年级"),
        db=None,
    ) == ["8年级"]
    assert get_accessible_grades(user("teacher"), db=None) == []


def test_add_in_condition_generates_explicit_placeholders():
    conditions = []
    params = {}

    added = _add_in_condition(conditions, params, "grade_level", ["7年级", "8年级"], "grade")

    assert added is True
    assert conditions == ["grade_level IN (:grade_0, :grade_1)"]
    assert params == {"grade_0": "7年级", "grade_1": "8年级"}


def test_class_contrast_uses_standard_library_statistics():
    rows = [
        SimpleNamespace(class_name="701", total_score=100),
        SimpleNamespace(class_name="701", total_score=80),
        SimpleNamespace(class_name="702", total_score=60),
        SimpleNamespace(class_name="702", total_score=40),
    ]

    result = analyze_class_contrast(rows, db=None)

    assert result["grade_mean"] == 70.0
    assert result["class_statistics"]["701"]["mean"] == 90.0
    assert result["class_statistics"]["702"]["mean"] == 50.0
    assert result["class_ranking"][0]["class_name"] == "701"


def test_analysis_workbook_exports_summary_and_structured_sheets():
    result = SimpleNamespace(
        analysis_id="analysis-1",
        exam_name="7年级期中考试",
        grade_level="7年级",
        analysis_type="overall",
        analysis_scope="all",
    )
    data = {
        "total_students": 2,
        "class_analysis": {
            "701班": {"mean": 90.0, "count": 1},
            "702班": {"mean": 80.0, "count": 1},
        },
        "top_improved": [
            {"student_name": "甲", "score_change": 12},
            {"student_name": "乙", "score_change": 8},
        ],
    }

    workbook_bytes = _build_analysis_workbook(result, data)
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)

    assert "分析概览" in workbook.sheetnames
    assert "class_analysis" in workbook.sheetnames
    assert "top_improved" in workbook.sheetnames
    assert workbook["分析概览"]["B2"].value == "analysis-1"
    assert workbook["class_analysis"]["A2"].value == "701班"
    assert workbook["top_improved"]["A1"].value == "student_name"
