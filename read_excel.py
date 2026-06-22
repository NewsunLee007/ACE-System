import openpyxl
import json

# 读取Excel文件
wb = openpyxl.load_workbook('/Users/newsunlee/Desktop/AI for learning/ACE System/2025-1 7年级期末考试-18班-成绩分析.xlsx')

# 获取所有sheet名称
print("Sheet names:", wb.sheetnames)

# 读取cs子表
if 'cs' in wb.sheetnames:
    ws = wb['cs']
    print("\n=== cs sheet content ===")
    
    # 读取前30行数据
    data = []
    for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
        data.append(row)
    
    for i, row in enumerate(data):
        print(f"Row {i+1}: {row}")

# 读取其他sheet
for sheet_name in wb.sheetnames[:3]:  # 只读前3个sheet
    if sheet_name != 'cs':
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} sheet (first 10 rows) ===")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
            print(f"Row {i+1}: {row}")
